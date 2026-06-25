"""
Alert Manager
=============
Creates and persists Alert objects for transactions at or above the
configured risk threshold. Handles deduplication within a time window.
"""

from __future__ import annotations

from dataclasses import asdict, dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import List

import pandas as pd

from src.utils.helpers import generate_alert_id, load_config, now_utc, ensure_dir
from src.utils.logger import get_logger

log = get_logger(__name__)


@dataclass
class Alert:
    """Represents a single AML alert for one transaction."""
    alert_id: str
    txn_id: str | int
    timestamp: str                    # ISO8601
    sender_account: str
    receiver_account: str
    amount: float
    payment_currency: str
    sender_location: str
    receiver_location: str
    payment_type: str
    risk_tier: str
    risk_score: int
    triggered_rules: str
    rule_reasons: str
    is_laundering_gt: int             # Ground truth label
    laundering_type_gt: str           # Ground truth laundering type
    created_at: str = field(default_factory=lambda: now_utc().isoformat())
    status: str = "OPEN"              # OPEN | UNDER_REVIEW | CLOSED | ESCALATED


class AlertManager:
    """
    Creates, deduplicates, and persists AML alerts.

    Usage:
        mgr = AlertManager()
        alerts = mgr.create_alerts(scored_df)
        mgr.save(alerts)
    """

    def __init__(self, config_path: str = "config/config.yaml"):
        cfg = load_config(config_path)
        alerts_cfg = cfg["alerts"]
        self.output_file = alerts_cfg["output_file"]
        self.excel_report = alerts_cfg["excel_report"]
        self.dedup_window = timedelta(hours=int(alerts_cfg.get("dedup_window_hours", 48)))

    # ── Public API ────────────────────────────────────────────────────────

    def create_alerts(self, df: pd.DataFrame) -> list[Alert]:
        """
        Generate Alert objects for all flagged transactions.

        Args:
            df: Output of RiskScorer.assign_tiers() with is_flagged column.

        Returns:
            List of Alert objects (deduplicated).
        """
        flagged = df[df["is_flagged"]].copy()
        log.info(f"Generating alerts for {len(flagged):,} flagged transactions …")

        alerts: list[Alert] = []
        seen: dict[str, datetime] = {}   # account → last alert time (for dedup)

        for idx, row in flagged.iterrows():
            acct = str(row["Sender_account"])
            ts = row.get("Timestamp")
            if pd.isna(ts):
                ts = None

            # Deduplication: skip if same account alerted within window
            if ts is not None and acct in seen:
                if (ts - seen[acct]) < self.dedup_window:
                    continue

            alert = Alert(
                alert_id=generate_alert_id(),
                txn_id=row.get("Txn_id", idx),
                timestamp=ts.isoformat() if ts and not pd.isna(ts) else "",
                sender_account=acct,
                receiver_account=str(row["Receiver_account"]),
                amount=float(row["Amount"]),
                payment_currency=str(row["Payment_currency"]),
                sender_location=str(row["Sender_bank_location"]),
                receiver_location=str(row["Receiver_bank_location"]),
                payment_type=str(row["Payment_type"]),
                risk_tier=str(row["risk_tier"]),
                risk_score=int(row["total_risk_score"]),
                triggered_rules=str(row.get("triggered_rules", "")),
                rule_reasons=str(row.get("rule_reasons", "")),
                is_laundering_gt=int(row.get("Is_laundering", 0)) if row.get("Is_laundering", -1) != -1 else 0,
                laundering_type_gt=str(row.get("Laundering_type", "")),
            )
            alerts.append(alert)
            if ts is not None:
                seen[acct] = ts

        log.info(f"Generated {len(alerts):,} alerts (after deduplication)")
        return alerts

    def save(self, alerts: list[Alert]) -> str:
        """Save alerts to CSV and Excel. Returns path to CSV."""
        if not alerts:
            log.warning("No alerts to save")
            return ""

        ensure_dir(Path(self.output_file).parent)
        ensure_dir(Path(self.excel_report).parent)

        alert_rows = [asdict(a) for a in alerts]
        df = pd.DataFrame(alert_rows)

        # CSV
        df.to_csv(self.output_file, index=False)
        log.info(f"Alerts saved → {self.output_file}  ({len(df):,} rows)")

        # Excel (colour-coded by risk tier)
        self._save_excel(df)
        return self.output_file

    def load_alerts(self) -> pd.DataFrame:
        """Load previously saved alerts CSV."""
        path = Path(self.output_file)
        if not path.exists():
            return pd.DataFrame()
        return pd.read_csv(path)

    def summary(self, alerts: list[Alert]) -> dict:
        """Return count breakdown by risk tier."""
        from collections import Counter
        counts = Counter(a.risk_tier for a in alerts)
        return dict(counts)

    # ── Private ───────────────────────────────────────────────────────────

    def _save_excel(self, df: pd.DataFrame) -> None:
        """Export colour-coded Excel alert report."""
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows

            tier_colors = {
                "MEDIUM":   "FFF9C4",  # light yellow
                "HIGH":     "FFB74D",  # orange
                "CRITICAL": "EF5350",  # red
            }

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "AML Alerts"

            # Header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # Style header
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Colour-code data rows by risk tier
            tier_col_idx = df.columns.get_loc("risk_tier") + 1  # 1-indexed
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                tier_cell = ws.cell(row=row_idx, column=tier_col_idx)
                tier_val = str(tier_cell.value)
                color = tier_colors.get(tier_val, "FFFFFF")
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for cell in row:
                    cell.fill = fill

            # Auto-fit column widths
            for col in ws.columns:
                max_len = max((len(str(c.value)) if c.value else 0) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

            wb.save(self.excel_report)
            log.info(f"Excel report saved → {self.excel_report}")
        except Exception as e:
            log.warning(f"Excel export failed: {e}")
